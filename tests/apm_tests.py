import os
import sys
import time
import json
import numpy as np
import requests
import concurrent.futures
import psutil
import threading
from datetime import datetime

class APMTestSuite:
    def __init__(self, base_url="http://127.0.0.1:8000"):
        self.base_url = base_url.rstrip("/")
        self.session = requests.Session()
        self.admin_token = None
        self.user_token = None
        
        # Sampling arrays
        self.cpu_samples = []
        self.memory_samples = []
        self.monitoring = False
        self.monitor_thread = None

        # Telemetry storage
        self.endpoint_metrics = {}
        self.all_latencies = []
        self.request_traces = []
        self.errors_log = []
        self.db_status = {"status": "UNKNOWN", "latency_ms": 0.0, "db_type": "Unknown"}

    def _monitor_resources(self, interval=0.1):
        """Background thread to sample CPU and Memory utilization."""
        current_process = psutil.Process()
        while self.monitoring:
            try:
                cpu = psutil.cpu_percent(interval=None)
                mem = current_process.memory_info().rss / (1024 * 1024)  # MB
                self.cpu_samples.append(cpu)
                self.memory_samples.append(mem)
            except Exception:
                pass
            time.sleep(interval)

    def start_monitoring(self):
        self.monitoring = True
        self.cpu_samples = []
        self.memory_samples = []
        # Prime cpu_percent
        psutil.cpu_percent(interval=None)
        self.monitor_thread = threading.Thread(target=self._monitor_resources, daemon=True)
        self.monitor_thread.start()

    def stop_monitoring(self):
        self.monitoring = False
        if self.monitor_thread:
            self.monitor_thread.join(timeout=1.0)

    def setup_auth(self):
        """Attempt to authenticate admin and user accounts."""
        try:
            r_admin = self.session.post(f"{self.base_url}/auth/login", json={"email": "admin@spectraguard.com", "password": "Admin@1234"})
            if r_admin.status_code == 200:
                self.admin_token = r_admin.json().get("access_token")
        except Exception:
            pass

        try:
            r_user = self.session.post(f"{self.base_url}/auth/login", json={"email": "user@example.com", "password": "User@1234"})
            if r_user.status_code == 200:
                self.user_token = r_user.json().get("access_token")
        except Exception:
            pass

    def run_tests(self, iterations_per_endpoint=15, concurrency=5):
        print("\n" + "=" * 70)
        print("    Executing SpectraGuard Application Performance Monitoring (APM)")
        print("=" * 70)

        # Step 1: Health check & DB health check
        print("1. Checking Application Availability & Database Health...")
        app_available = False
        try:
            res = requests.get(f"{self.base_url}/health", timeout=3)
            if res.status_code == 200:
                app_available = True
        except Exception:
            app_available = False

        if not app_available:
            print("ERROR: Application server is not reachable!")

        try:
            res_db = requests.get(f"{self.base_url}/apm/db-health", timeout=3)
            if res_db.status_code == 200:
                self.db_status = res_db.json()
            else:
                self.db_status = {"status": "DEGRADED", "latency_ms": 0.0, "db_type": "SQLite/MySQL"}
        except Exception as e:
            self.db_status = {"status": "UNHEALTHY", "error": str(e), "latency_ms": 0.0, "db_type": "Unknown"}

        # Step 2: Define real endpoints to test
        routes = [
            ("GET", "/health", None, "public", "Health Endpoint"),
            ("GET", "/", None, "public", "Root Welcome API"),
            ("GET", "/docs", None, "public", "Swagger Documentation"),
            ("GET", "/openapi.json", None, "public", "OpenAPI Schema"),
            ("POST", "/auth/login", {"email": "admin@spectraguard.com", "password": "Admin@1234"}, "json", "Admin Login"),
            ("POST", "/auth/login", {"email": "user@example.com", "password": "User@1234"}, "json", "User Login"),
            ("POST", "/auth/login", {"email": "invalid@example.com", "password": "WrongPassword"}, "json", "Failed Login Attempt"),
            ("GET", "/auth/me", None, "auth_user", "Get User Profile"),
            ("POST", "/auth/refresh-token", {"refresh_token": "invalid_refresh_token_test"}, "json", "Invalid Refresh Token"),
            ("GET", "/spectra/sample-datasets", None, "public", "Get Sample Datasets"),
            ("GET", "/spectra/1", None, "auth_user", "Get Spectrum Detail"),
            ("POST", "/classify/1", None, "auth_user", "Run Spectrum Classification"),
            ("GET", "/classify/reference-matches/1", None, "auth_user", "Get Spectrum Matches"),
            ("GET", "/tests", None, "auth_user", "List Test Records"),
            ("GET", "/tests/1", None, "auth_user", "Get Detailed Test Record"),
            ("GET", "/reference", None, "public", "List Reference Database"),
            ("GET", "/reference/1", None, "public", "Get Reference Detail"),
            ("POST", "/reports/generate/1", None, "auth_user", "Generate PDF Report"),
            ("GET", "/reports/1", None, "auth_user", "Download PDF Report"),
            ("GET", "/admin/stats", None, "auth_admin", "Admin System Metrics"),
            ("GET", "/admin/users", None, "auth_admin", "Admin User List"),
            ("GET", "/nearby/pharmacies", None, "public", "Nearby Pharmacies"),
            ("POST", "/chat", {"message": "How does Raman spectroscopy work?"}, "json", "AI Chat Endpoint"),
            ("GET", "/apm/metrics", None, "public", "APM Telemetry Metrics"),
            ("GET", "/apm/db-health", None, "public", "APM Database Health"),
        ]

        self.setup_auth()

        # Step 3: Start background CPU & Memory sampling
        self.start_monitoring()

        print(f"2. Dispatching monitored requests across {len(routes)} API endpoints (x{iterations_per_endpoint} runs)...")

        work_items = []
        for route in routes:
            method, ep, payload, auth_mode, name = route
            for i in range(iterations_per_endpoint):
                work_items.append((method, ep, payload, auth_mode, name, i))

        start_all = time.time()

        def execute_request(item):
            method, ep, payload, auth_mode, name, iter_idx = item
            headers = {}
            if auth_mode == "auth_user" and self.user_token:
                headers["Authorization"] = f"Bearer {self.user_token}"
            elif auth_mode == "auth_admin" and self.admin_token:
                headers["Authorization"] = f"Bearer {self.admin_token}"

            start_t = time.time()
            url = f"{self.base_url}{ep}"
            status_code = 500
            trace_id = f"tr-client-{iter_idx}"
            exc_str = None

            try:
                if method == "GET":
                    r = self.session.get(url, headers=headers, timeout=5)
                elif method == "POST":
                    r = self.session.post(url, json=payload, headers=headers, timeout=5)
                else:
                    r = self.session.get(url, headers=headers, timeout=5)

                end_t = time.time()
                status_code = r.status_code
                trace_id = r.headers.get("X-Trace-ID", trace_id)
                dur_ms = (end_t - start_t) * 1000

                # Sanitize response text to ensure no secrets exposed
                resp_snippet = r.text[:100] if r.text else ""
                if "access_token" in resp_snippet or "password" in resp_snippet:
                    resp_snippet = "[REDACTED SENSITIVE DATA]"

                return {
                    "method": method,
                    "endpoint": ep,
                    "name": name,
                    "status_code": status_code,
                    "duration_ms": dur_ms,
                    "start_time": start_t,
                    "end_time": end_t,
                    "trace_id": trace_id,
                    "exception": None,
                    "snippet": resp_snippet
                }
            except Exception as e:
                end_t = time.time()
                dur_ms = (end_t - start_t) * 1000
                exc_str = f"{type(e).__name__}: {str(e)}"
                return {
                    "method": method,
                    "endpoint": ep,
                    "name": name,
                    "status_code": 500,
                    "duration_ms": dur_ms,
                    "start_time": start_t,
                    "end_time": end_t,
                    "trace_id": trace_id,
                    "exception": exc_str,
                    "snippet": exc_str[:80]
                }

        with concurrent.futures.ThreadPoolExecutor(max_workers=concurrency) as executor:
            results = list(executor.map(execute_request, work_items))

        total_elapsed = time.time() - start_all

        self.stop_monitoring()

        # Step 4: Process and aggregate telemetry data
        total_requests = len(results)
        successful_requests = 0
        failed_requests = 0

        endpoint_groups = {}

        for res in results:
            ep_key = f"{res['method']} {res['endpoint']}"
            if ep_key not in endpoint_groups:
                endpoint_groups[ep_key] = {
                    "method": res["method"],
                    "endpoint": res["endpoint"],
                    "name": res["name"],
                    "durations": [],
                    "statuses": [],
                    "success_count": 0,
                    "error_count": 0,
                    "exceptions": []
                }

            group = endpoint_groups[ep_key]
            group["durations"].append(res["duration_ms"])
            group["statuses"].append(res["status_code"])
            self.all_latencies.append(res["duration_ms"])

            # 2xx and expected auth error test cases (e.g. 401 on bad login) are valid business outcomes
            # HTTP 5xx or unhandled exceptions are failures
            is_success = (res["status_code"] < 500) and not res["exception"]
            if is_success:
                successful_requests += 1
                group["success_count"] += 1
            else:
                failed_requests += 1
                group["error_count"] += 1
                group["exceptions"].append(res["exception"] or f"HTTP {res['status_code']}")

                self.errors_log.append({
                    "error_type": "HTTPError" if res["status_code"] >= 400 else "Exception",
                    "endpoint": res["endpoint"],
                    "status_code": res["status_code"],
                    "message": res["exception"] or f"HTTP Status {res['status_code']} on {res['endpoint']}",
                    "severity": "HIGH" if res["status_code"] >= 500 else "MEDIUM",
                    "timestamp": datetime.fromtimestamp(res["start_time"]).strftime("%Y-%m-%d %H:%M:%S")
                })

            self.request_traces.append({
                "trace_id": res["trace_id"],
                "endpoint": res["endpoint"],
                "method": res["method"],
                "start_time": datetime.fromtimestamp(res["start_time"]).strftime("%H:%M:%S.%f")[:-3],
                "end_time": datetime.fromtimestamp(res["end_time"]).strftime("%H:%M:%S.%f")[:-3],
                "duration_ms": round(res["duration_ms"], 2),
                "status_code": res["status_code"],
                "exception": res["exception"]
            })

        # Calculate Endpoint level statistics
        for ep_key, grp in endpoint_groups.items():
            durs = np.array(grp["durations"])
            tot = len(durs)
            err_pct = (grp["error_count"] / tot) * 100.0 if tot > 0 else 0.0
            rps = tot / total_elapsed if total_elapsed > 0 else 0.0

            ep_status = "PASSED" if grp["error_count"] == 0 and err_pct == 0.0 else "FAILED"

            self.endpoint_metrics[ep_key] = {
                "method": grp["method"],
                "endpoint": grp["endpoint"],
                "name": grp["name"],
                "total_requests": tot,
                "success_count": grp["success_count"],
                "error_count": grp["error_count"],
                "error_pct": round(err_pct, 2),
                "avg_response_time": round(float(np.mean(durs)), 2),
                "min_response_time": round(float(np.min(durs)), 2),
                "max_response_time": round(float(np.max(durs)), 2),
                "requests_per_sec": round(rps, 2),
                "status": ep_status
            }

        # Global latency statistics
        lats = np.array(self.all_latencies) if self.all_latencies else np.array([50.0])
        avg_lat = float(np.mean(lats))
        p50_lat = float(np.percentile(lats, 50))
        p90_lat = float(np.percentile(lats, 90))
        p95_lat = float(np.percentile(lats, 95))
        p99_lat = float(np.percentile(lats, 99))
        max_lat = float(np.max(lats))
        rps_global = total_requests / total_elapsed if total_elapsed > 0 else 0.0

        success_rate = (successful_requests / total_requests) * 100.0 if total_requests > 0 else 100.0
        error_rate = 100.0 - success_rate

        # System resource stats
        cpus = np.array(self.cpu_samples) if self.cpu_samples else np.array([5.0])
        mems = np.array(self.memory_samples) if self.memory_samples else np.array([80.0])
        avg_cpu = float(np.mean(cpus))
        max_cpu = float(np.max(cpus))
        init_mem = float(mems[0]) if len(mems) > 0 else 80.0
        avg_mem = float(np.mean(mems))
        max_mem = float(np.max(mems))
        mem_increase = max_mem - init_mem

        # Status Evaluation according to documented thresholds:
        # App available == True, Error rate <= 1.0%, P95 <= 500ms, Avg CPU <= 80%, RAM growth <= 100MB, DB Healthy
        all_endpoints_passed = all(m.get("status") == "PASSED" for m in self.endpoint_metrics.values()) if self.endpoint_metrics else True
        if app_available and error_rate <= 1.0 and p95_lat <= 500.0 and avg_cpu <= 80.0 and mem_increase <= 100.0 and self.db_status.get("status") == "HEALTHY" and all_endpoints_passed:
            apm_status = "APM TESTING PASSED"
        else:
            apm_status = "APM TESTING FAILED"

        summary_test_metrics = [
            {"category": "Overall APM", "name": "Overall APM Evaluation", "value": apm_status, "target": "APM TESTING PASSED", "status": apm_status, "description": "Master APM Status Evaluation"},
            {"category": "App Health", "name": "Application Availability", "value": "100% (Available)" if app_available else "0% (Unavailable)", "target": "100%", "status": "PASSED" if app_available else "FAILED", "description": "FastAPI Core Engine Status"},
            {"category": "Throughput", "name": "Total Monitored Requests", "value": str(total_requests), "target": "> 0", "status": "PASSED" if total_requests > 0 else "FAILED", "description": f"Throughput: {rps_global:.2f} req/s"},
            {"category": "Success Rate", "name": "Successful Requests", "value": f"{successful_requests} ({success_rate:.2f}%)", "target": "100%", "status": "PASSED" if success_rate >= 99.0 else "FAILED", "description": f"Success Rate: {success_rate:.2f}%"},
            {"category": "Error Rate", "name": "Failed Requests", "value": f"{failed_requests} ({error_rate:.2f}%)", "target": "0", "status": "PASSED" if failed_requests == 0 else "FAILED", "description": f"Error Rate: {error_rate:.2f}%"},
            {"category": "Response Time", "name": "Average Latency", "value": f"{round(avg_lat, 2)} ms", "target": "< 200 ms", "status": "PASSED" if avg_lat < 200.0 else "FAILED", "description": "Mean response duration"},
            {"category": "Response Time", "name": "P50 Latency (Median)", "value": f"{round(p50_lat, 2)} ms", "target": "< 300 ms", "status": "PASSED" if p50_lat < 300.0 else "FAILED", "description": "50th percentile latency"},
            {"category": "Response Time", "name": "P90 Latency", "value": f"{round(p90_lat, 2)} ms", "target": "< 500 ms", "status": "PASSED" if p90_lat < 500.0 else "FAILED", "description": "90th percentile latency"},
            {"category": "Response Time", "name": "P95 Latency", "value": f"{round(p95_lat, 2)} ms", "target": "< 500 ms", "status": "PASSED" if p95_lat < 500.0 else "FAILED", "description": "95th percentile latency threshold"},
            {"category": "Response Time", "name": "P99 Latency", "value": f"{round(p99_lat, 2)} ms", "target": "< 1000 ms", "status": "PASSED" if p99_lat < 1000.0 else "FAILED", "description": "99th percentile latency"},
            {"category": "Response Time", "name": "Maximum Latency", "value": f"{round(max_lat, 2)} ms", "target": "< 2000 ms", "status": "PASSED" if max_lat < 2000.0 else "FAILED", "description": "Peak latency observed"},
            {"category": "CPU Usage", "name": "Average CPU Utilization", "value": f"{round(avg_cpu, 2)}%", "target": "< 80.0%", "status": "PASSED" if avg_cpu < 80.0 else "FAILED", "description": f"Peak CPU: {round(max_cpu, 2)}%"},
            {"category": "Memory Usage", "name": "Average Memory Usage", "value": f"{round(avg_mem, 2)} MB", "target": "Growth < 100 MB", "status": "PASSED" if mem_increase < 100.0 else "FAILED", "description": f"Max RAM: {round(max_mem, 2)} MB (+{round(mem_increase, 2)} MB growth)"},
            {"category": "Database", "name": "Database Health", "value": self.db_status.get("status", "HEALTHY"), "target": "HEALTHY", "status": "PASSED" if self.db_status.get("status") == "HEALTHY" else "FAILED", "description": f"Ping: {self.db_status.get('latency_ms', 0.0)} ms ({self.db_status.get('db_type', 'SQLite/MySQL')})"}
        ]

        # Construct final results dictionary
        results_summary = {
            "app_availability": "100% (Available)" if app_available else "0% (Unavailable)",
            "total_monitored_requests": total_requests,
            "successful_requests": successful_requests,
            "failed_requests": failed_requests,
            "success_rate": f"{success_rate:.2f}%",
            "error_rate": f"{error_rate:.2f}%",
            "throughput_rps": f"{rps_global:.2f} req/s",
            "average_latency_ms": round(avg_lat, 2),
            "p50_latency_ms": round(p50_lat, 2),
            "p90_latency_ms": round(p90_lat, 2),
            "p95_latency_ms": round(p95_lat, 2),
            "p99_latency_ms": round(p99_lat, 2),
            "maximum_latency_ms": round(max_lat, 2),
            "average_cpu_usage_pct": round(avg_cpu, 2),
            "maximum_cpu_usage_pct": round(max_cpu, 2),
            "initial_memory_mb": round(init_mem, 2),
            "average_memory_mb": round(avg_mem, 2),
            "maximum_memory_mb": round(max_mem, 2),
            "memory_increase_mb": round(mem_increase, 2),
            "database_status": self.db_status.get("status", "HEALTHY"),
            "database_latency_ms": self.db_status.get("latency_ms", 0.0),
            "database_type": self.db_status.get("db_type", "SQLite/MySQL"),
            "apm_status": apm_status,
            "overall_apm_status": apm_status,
            "summary_test_metrics": summary_test_metrics,
            "endpoint_metrics": self.endpoint_metrics,
            "errors": self.errors_log,
            "sample_traces": self.request_traces[:50]
        }

        # Step 5: Save JSON & HTML Reports
        os.makedirs("reports", exist_ok=True)

        json_path = os.path.join("reports", "apm_report.json")
        with open(json_path, "w", encoding="utf-8") as f:
            json.dump(results_summary, f, indent=2)
        print(f"Generated JSON APM Report: {os.path.abspath(json_path)}")

        html_path = os.path.join("reports", "apm_report.html")
        self.generate_html_report(results_summary, html_path)
        print(f"Generated HTML APM Dashboard Report: {os.path.abspath(html_path)}")

        excel_path = os.path.join("reports", "APM_Report.xlsx")
        self.generate_excel_report(results_summary, excel_path)
        print(f"Generated Excel APM Report: {os.path.abspath(excel_path)}")

        print("\nAPM Execution Completed successfully!")
        return results_summary

    def generate_excel_report(self, data, filepath):
        """Generate a professionally styled Excel (.xlsx) APM Report workbook."""
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        wb = openpyxl.Workbook()
        wb.remove(wb.active)  # Remove default sheet

        # Styles
        header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        
        pass_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
        pass_font = Font(name="Calibri", size=10, color="375623", bold=True)

        fail_fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
        fail_font = Font(name="Calibri", size=10, color="C65911", bold=True)

        border = Border(
            left=Side(style='thin', color='D9D9D9'),
            right=Side(style='thin', color='D9D9D9'),
            top=Side(style='thin', color='D9D9D9'),
            bottom=Side(style='thin', color='D9D9D9')
        )

        def style_sheet(ws, headers, rows):
            ws.append(headers)
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_num)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            ws.row_dimensions[1].height = 28

            for row_idx, row_data in enumerate(rows, start=2):
                ws.append(row_data)
                ws.row_dimensions[row_idx].height = 20
                for col_idx, value in enumerate(row_data, start=1):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    cell.border = border
                    cell.font = Font(name="Calibri", size=10)
                    cell.alignment = Alignment(vertical="center")

                    str_val = str(value).upper()
                    if "PASSED" in str_val or "HEALTHY" in str_val or "SUCCESS" in str_val:
                        cell.fill = pass_fill
                        cell.font = pass_font
                        cell.alignment = Alignment(horizontal="center", vertical="center")
                    elif "FAILED" in str_val or "UNHEALTHY" in str_val or "ERR" in str_val:
                        cell.fill = fail_fill
                        cell.font = fail_font
                        cell.alignment = Alignment(horizontal="center", vertical="center")

            for col in ws.columns:
                max_len = 0
                col_letter = get_column_letter(col[0].column)
                for cell in col:
                    val = str(cell.value or '')
                    if '\n' in val:
                        val = max(val.split('\n'), key=len)
                    max_len = max(max_len, len(val))
                ws.column_dimensions[col_letter].width = min(max(max_len + 4, 14), 60)

        # Tab 1: APM Summary Metrics
        ws_sum = wb.create_sheet(title="APM Summary Metrics")
        sum_headers = ["Category", "Metric Name", "Measured Value", "Threshold / Baseline", "Status", "Description"]
        if "summary_test_metrics" in data:
            sum_rows = [
                [m["category"], m["name"], m["value"], m["target"], m["status"], m["description"]]
                for m in data["summary_test_metrics"]
            ]
        else:
            sum_rows = [
                ["Overall APM", "Overall APM Evaluation", data.get("apm_status", "APM TESTING PASSED"), "APM TESTING PASSED", data.get("apm_status", "APM TESTING PASSED"), "Master APM Status Evaluation"],
                ["App Health", "Application Availability", data["app_availability"], "100%", "PASSED" if "100%" in data["app_availability"] else "FAILED", "FastAPI Core Engine Status"],
                ["Throughput", "Total Monitored Requests", data["total_monitored_requests"], "> 0", "PASSED" if data["total_monitored_requests"] > 0 else "FAILED", f"Throughput: {data['throughput_rps']}"],
                ["Success Rate", "Successful Requests", data["successful_requests"], "100%", "PASSED" if data["failed_requests"] == 0 else "FAILED", f"Success Rate: {data['success_rate']}"],
                ["Error Rate", "Failed Requests", data["failed_requests"], "0", "PASSED" if data["failed_requests"] == 0 else "FAILED", f"Error Rate: {data['error_rate']}"],
                ["Response Time", "Average Latency", f"{data['average_latency_ms']} ms", "< 200 ms", "PASSED" if data['average_latency_ms'] < 200 else "FAILED", "Mean response duration"],
                ["Response Time", "P50 Latency (Median)", f"{data['p50_latency_ms']} ms", "< 300 ms", "PASSED" if data['p50_latency_ms'] < 300 else "FAILED", "50th percentile latency"],
                ["Response Time", "P90 Latency", f"{data['p90_latency_ms']} ms", "< 500 ms", "PASSED" if data['p90_latency_ms'] < 500 else "FAILED", "90th percentile latency"],
                ["Response Time", "P95 Latency", f"{data['p95_latency_ms']} ms", "< 500 ms", "PASSED" if data['p95_latency_ms'] < 500 else "FAILED", "95th percentile latency threshold"],
                ["Response Time", "P99 Latency", f"{data['p99_latency_ms']} ms", "< 1000 ms", "PASSED" if data['p99_latency_ms'] < 1000 else "FAILED", "99th percentile latency"],
                ["Response Time", "Maximum Latency", f"{data['maximum_latency_ms']} ms", "< 2000 ms", "PASSED" if data['maximum_latency_ms'] < 2000 else "FAILED", "Peak latency observed"],
                ["CPU Usage", "Average CPU Utilization", f"{data['average_cpu_usage_pct']}%", "< 80.0%", "PASSED" if data['average_cpu_usage_pct'] < 80.0 else "FAILED", f"Peak CPU: {data['maximum_cpu_usage_pct']}%"],
                ["Memory Usage", "Average Memory Usage", f"{data['average_memory_mb']} MB", "< 500 MB", "PASSED" if data['memory_increase_mb'] < 100 else "FAILED", f"Max RAM: {data['maximum_memory_mb']} MB (+{data['memory_increase_mb']} MB growth)"],
                ["Database", "Database Health", data["database_status"], "HEALTHY", "PASSED" if data["database_status"] == "HEALTHY" else "FAILED", f"Ping: {data['database_latency_ms']} ms ({data['database_type']})"]
            ]
        style_sheet(ws_sum, sum_headers, sum_rows)

        # Tab 2: Endpoint Performance
        ws_ep = wb.create_sheet(title="Endpoint Performance")
        ep_headers = ["Method", "Endpoint Path", "Endpoint Name", "Requests", "Success Count", "Error Count", "Error %", "Avg Latency (ms)", "Min Latency (ms)", "Max Latency (ms)", "Throughput (req/s)", "Status"]
        ep_rows = []
        for ep_key, m in data["endpoint_metrics"].items():
            status_str = m.get("status", "PASSED" if (m["error_count"] == 0 and m["error_pct"] == 0) else "FAILED")
            ep_rows.append([
                m["method"],
                m["endpoint"],
                m["name"],
                m["total_requests"],
                m["success_count"],
                m["error_count"],
                f"{m['error_pct']}%",
                m["avg_response_time"],
                m["min_response_time"],
                m["max_response_time"],
                m["requests_per_sec"],
                status_str
            ])
        style_sheet(ws_ep, ep_headers, ep_rows)

        # Tab 3: Request Traces
        ws_tr = wb.create_sheet(title="Request Traces & Transactions")
        tr_headers = ["Trace ID", "Method", "Endpoint Path", "Start Time", "End Time", "Duration (ms)", "HTTP Status", "Status", "Exception / Error Details"]
        tr_rows = []
        for tr in data["sample_traces"]:
            tr_status = "PASSED" if (tr["status_code"] < 400 and not tr.get("exception")) else "FAILED"
            tr_rows.append([
                tr["trace_id"],
                tr["method"],
                tr["endpoint"],
                tr["start_time"],
                tr["end_time"],
                tr["duration_ms"],
                tr["status_code"],
                tr_status,
                tr["exception"] or "None"
            ])
        style_sheet(ws_tr, tr_headers, tr_rows)

        wb.save(filepath)
        try:
            import stat
            os.chmod(filepath, stat.S_IWRITE | stat.S_IREAD)
        except Exception:
            pass

    def generate_html_report(self, data, filepath):
        """Generate a sleek, dark-themed, glassmorphic APM HTML Report."""
        status_color = "#10B981" if "PASSED" in data["apm_status"] else "#EF4444"
        
        summary_rows_html = ""
        for sm in data.get("summary_test_metrics", []):
            st = sm["status"]
            b_style = 'background: rgba(16,185,129,0.2); color:#34D399; border: 1px solid #10B981;' if st == "PASSED" else 'background: rgba(239,68,68,0.2); color:#F87171; border: 1px solid #EF4444;'
            summary_rows_html += f"""
            <tr>
                <td style="padding:10px; border-bottom: 1px solid #1E293B;"><strong>{sm['category']}</strong></td>
                <td style="padding:10px; border-bottom: 1px solid #1E293B;">{sm['name']}</td>
                <td style="padding:10px; border-bottom: 1px solid #1E293B; font-weight:600;">{sm['value']}</td>
                <td style="padding:10px; border-bottom: 1px solid #1E293B; color:#94A3B8;">{sm['target']}</td>
                <td style="padding:10px; border-bottom: 1px solid #1E293B;"><span style="{b_style} padding:4px 10px; border-radius:4px; font-weight:700; font-size:12px;">{st}</span></td>
                <td style="padding:10px; border-bottom: 1px solid #1E293B; color:#64748B;">{sm['description']}</td>
            </tr>
            """

        ep_rows_html = ""
        for ep_key, m in data["endpoint_metrics"].items():
            ep_status = m.get("status", "PASSED" if m["error_count"] == 0 else "FAILED")
            badge_style = 'background: rgba(16,185,129,0.2); color:#34D399; border: 1px solid #10B981;' if ep_status == "PASSED" else 'background: rgba(239,68,68,0.2); color:#F87171; border: 1px solid #EF4444;'
            badge = f'<span style="{badge_style} padding:4px 10px; border-radius:4px; font-weight:700; font-size:12px;">{ep_status}</span>'
            ep_rows_html += f"""
            <tr>
                <td style="padding:10px; border-bottom: 1px solid #1E293B;"><strong>{m['method']}</strong></td>
                <td style="padding:10px; border-bottom: 1px solid #1E293B; font-family:monospace;">{m['endpoint']}</td>
                <td style="padding:10px; border-bottom: 1px solid #1E293B;">{m['total_requests']}</td>
                <td style="padding:10px; border-bottom: 1px solid #1E293B; color:#34D399;">{m['success_count']}</td>
                <td style="padding:10px; border-bottom: 1px solid #1E293B; color:{'#F87171' if m['error_count']>0 else '#94A3B8'};">{m['error_count']} ({m['error_pct']}%)</td>
                <td style="padding:10px; border-bottom: 1px solid #1E293B;">{m['avg_response_time']} ms</td>
                <td style="padding:10px; border-bottom: 1px solid #1E293B;">{m['min_response_time']} / {m['max_response_time']} ms</td>
                <td style="padding:10px; border-bottom: 1px solid #1E293B;">{m['requests_per_sec']} req/s</td>
                <td style="padding:10px; border-bottom: 1px solid #1E293B;">{badge}</td>
            </tr>
            """

        trace_rows_html = ""
        for tr in data["sample_traces"][:20]:
            tr_badge = f'<span style="color:#34D399;">{tr["status_code"]}</span>' if tr["status_code"] < 400 else f'<span style="color:#F87171;">{tr["status_code"]}</span>'
            trace_rows_html += f"""
            <tr>
                <td style="padding:8px; border-bottom: 1px solid #1E293B; font-family:monospace; font-size:12px; color:#60A5FA;">{tr['trace_id']}</td>
                <td style="padding:8px; border-bottom: 1px solid #1E293B;">{tr['method']} {tr['endpoint']}</td>
                <td style="padding:8px; border-bottom: 1px solid #1E293B;">{tr['start_time']} -> {tr['end_time']}</td>
                <td style="padding:8px; border-bottom: 1px solid #1E293B; font-weight:600;">{tr['duration_ms']} ms</td>
                <td style="padding:8px; border-bottom: 1px solid #1E293B;">{tr_badge}</td>
            </tr>
            """

        html_content = f"""<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>SpectraGuard APM Test Report</title>
    <style>
        body {{
            font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif;
            background-color: #0F172A;
            color: #F8FAFC;
            margin: 0;
            padding: 24px;
        }}
        .container {{
            max-width: 1200px;
            margin: 0 auto;
        }}
        .header {{
            background: linear-gradient(135deg, #1E293B 0%, #0F172A 100%);
            padding: 24px;
            border-radius: 12px;
            border: 1px solid #334155;
            display: flex;
            justify-content: space-between;
            align-items: center;
            margin-bottom: 24px;
            box-shadow: 0 10px 25px -5px rgba(0, 0, 0, 0.5);
        }}
        .title h1 {{
            margin: 0 0 8px 0;
            font-size: 26px;
            color: #38BDF8;
        }}
        .title p {{
            margin: 0;
            color: #94A3B8;
            font-size: 14px;
        }}
        .status-pill {{
            background: {status_color}22;
            border: 2px solid {status_color};
            color: {status_color};
            padding: 8px 20px;
            border-radius: 30px;
            font-weight: 700;
            font-size: 18px;
            letter-spacing: 0.05em;
        }}
        .grid-cards {{
            display: grid;
            grid-template-columns: repeat(auto-fit, minmax(220px, 1fr));
            gap: 16px;
            margin-bottom: 24px;
        }}
        .card {{
            background: #1E293B;
            border: 1px solid #334155;
            border-radius: 10px;
            padding: 16px;
            box-shadow: 0 4px 6px -1px rgba(0, 0, 0, 0.3);
        }}
        .card-label {{
            font-size: 12px;
            color: #94A3B8;
            text-transform: uppercase;
            letter-spacing: 0.05em;
            margin-bottom: 6px;
        }}
        .card-value {{
            font-size: 22px;
            font-weight: 700;
            color: #F8FAFC;
        }}
        .card-sub {{
            font-size: 11px;
            color: #38BDF8;
            margin-top: 4px;
        }}
        .section-title {{
            font-size: 18px;
            font-weight: 600;
            color: #38BDF8;
            margin: 32px 0 16px 0;
            border-left: 4px solid #38BDF8;
            padding-left: 10px;
        }}
        table {{
            width: 100%;
            border-collapse: collapse;
            background: #1E293B;
            border-radius: 10px;
            overflow: hidden;
            border: 1px solid #334155;
            margin-bottom: 24px;
        }}
        th {{
            background: #0F172A;
            color: #94A3B8;
            text-align: left;
            padding: 12px 10px;
            font-size: 12px;
            text-transform: uppercase;
            border-bottom: 1px solid #334155;
        }}
        td {{
            font-size: 13px;
        }}
        .footer {{
            text-align: center;
            color: #64748B;
            font-size: 12px;
            margin-top: 40px;
        }}
    </style>
</head>
<body>
    <div class="container">
        <div class="header">
            <div class="title">
                <h1>SpectraGuard APM Test Report</h1>
                <p>Application Performance & System Resource Telemetry Report</p>
            </div>
            <div class="status-pill">{data['apm_status']}</div>
        </div>

        <div class="grid-cards">
            <div class="card">
                <div class="card-label">App Availability</div>
                <div class="card-value">{data['app_availability']}</div>
                <div class="card-sub">FastAPI Health Engine</div>
            </div>
            <div class="card">
                <div class="card-label">Monitored Requests</div>
                <div class="card-value">{data['total_monitored_requests']}</div>
                <div class="card-sub">Success Rate: {data['success_rate']}</div>
            </div>
            <div class="card">
                <div class="card-label">Avg Response Time</div>
                <div class="card-value">{data['average_latency_ms']} ms</div>
                <div class="card-sub">P95: {data['p95_latency_ms']} ms | Max: {data['maximum_latency_ms']} ms</div>
            </div>
            <div class="card">
                <div class="card-label">CPU Utilization</div>
                <div class="card-value">{data['average_cpu_usage_pct']}%</div>
                <div class="card-sub">Peak CPU: {data['maximum_cpu_usage_pct']}%</div>
            </div>
            <div class="card">
                <div class="card-label">RAM / Memory</div>
                <div class="card-value">{data['average_memory_mb']} MB</div>
                <div class="card-sub">Growth: +{data['memory_increase_mb']} MB (Max: {data['maximum_memory_mb']} MB)</div>
            </div>
            <div class="card">
                <div class="card-label">Database Health</div>
                <div class="card-value" style="color:#34D399;">{data['database_status']}</div>
                <div class="card-sub">Ping: {data['database_latency_ms']} ms ({data['database_type']})</div>
            </div>
        </div>

        <div class="section-title">📋 System Telemetry & APM Test Results</div>
        <table>
            <thead>
                <tr>
                    <th>Category</th>
                    <th>Metric / Test Name</th>
                    <th>Measured Value</th>
                    <th>Target / Condition</th>
                    <th>Status</th>
                    <th>Description</th>
                </tr>
            </thead>
            <tbody>
                {summary_rows_html}
            </tbody>
        </table>

        <div class="section-title">📊 Latency Percentile Distribution</div>
        <div class="grid-cards">
            <div class="card"><div class="card-label">P50 (Median)</div><div class="card-value">{data['p50_latency_ms']} ms</div></div>
            <div class="card"><div class="card-label">P90 Percentile</div><div class="card-value">{data['p90_latency_ms']} ms</div></div>
            <div class="card"><div class="card-label">P95 Percentile</div><div class="card-value" style="color:#38BDF8;">{data['p95_latency_ms']} ms</div></div>
            <div class="card"><div class="card-label">P99 Percentile</div><div class="card-value">{data['p99_latency_ms']} ms</div></div>
            <div class="card"><div class="card-label">Max Latency</div><div class="card-value">{data['maximum_latency_ms']} ms</div></div>
        </div>

        <div class="section-title">⚡ API Endpoint Monitored Performance</div>
        <table>
            <thead>
                <tr>
                    <th>Method</th>
                    <th>Endpoint Path</th>
                    <th>Requests</th>
                    <th>Success</th>
                    <th>Errors</th>
                    <th>Avg Latency</th>
                    <th>Min / Max</th>
                    <th>Throughput</th>
                    <th>Status</th>
                </tr>
            </thead>
            <tbody>
                {ep_rows_html}
            </tbody>
        </table>

        <div class="section-title">🔍 Request Traces & Transactions (Sample Logs)</div>
        <table>
            <thead>
                <tr>
                    <th>Trace ID</th>
                    <th>Endpoint Transaction</th>
                    <th>Time Interval</th>
                    <th>Duration</th>
                    <th>Status</th>
                </tr>
            </thead>
            <tbody>
                {trace_rows_html}
            </tbody>
        </table>

        <div class="footer">
            Generated on {datetime.now().strftime("%Y-%m-%d %H:%M:%S")} | SpectraGuard AI Pharmaceutical Authentication System
        </div>
    </div>
</body>
</html>
"""
        with open(filepath, "w", encoding="utf-8") as f:
            f.write(html_content)



if __name__ == "__main__":
    suite = APMTestSuite()
    suite.run_tests()

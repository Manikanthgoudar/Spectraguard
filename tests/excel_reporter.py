import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

class ExcelReporter:
    """
    Generates a professionally formatted multi-tab Excel workbook containing
    all test case results for Selenium E2E, API Integration, Load Testing, and Vulnerability Scanning.
    """
    def __init__(self, filepath="SpectraGuard_Test_Report.xlsx"):
        self.filepath = filepath
        self.wb = openpyxl.Workbook()
        # Remove default sheet
        self.wb.remove(self.wb.active)

        # Header style settings
        self.header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        self.header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        
        self.pass_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
        self.pass_font = Font(name="Calibri", size=10, color="375623", bold=True)

        self.fail_fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
        self.fail_font = Font(name="Calibri", size=10, color="C65911", bold=True)

        self.border = Border(
            left=Side(style='thin', color='D9D9D9'),
            right=Side(style='thin', color='D9D9D9'),
            top=Side(style='thin', color='D9D9D9'),
            bottom=Side(style='thin', color='D9D9D9')
        )

    def add_tab(self, title, headers, rows):
        ws = self.wb.create_sheet(title=title)
        
        # Write headers
        ws.append(headers)
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        ws.row_dimensions[1].height = 28

        # Write data rows
        for row_idx, row_data in enumerate(rows, start=2):
            ws.append(row_data)
            row_dim_height = 20
            ws.row_dimensions[row_idx].height = row_dim_height

            for col_idx, value in enumerate(row_data, start=1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.border = self.border
                cell.font = Font(name="Calibri", size=10)
                cell.alignment = Alignment(vertical="center")

                # Format Status column if present
                str_val = str(value).upper()
                if str_val in ["PASSED", "PASS", "GREEN", "SUCCESS"]:
                    cell.fill = self.pass_fill
                    cell.font = self.pass_font
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                elif str_val in ["FAILED", "FAIL", "RED", "VULNERABLE"]:
                    cell.fill = self.fail_fill
                    cell.font = self.fail_font
                    cell.alignment = Alignment(horizontal="center", vertical="center")

        # Auto-adjust column widths
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                val = str(cell.value or '')
                if '\n' in val:
                    val = max(val.split('\n'), key=len)
                max_len = max(max_len, len(val))
            ws.column_dimensions[col_letter].width = min(max(max_len + 4, 12), 60)

    def save(self):
        self.wb.save(self.filepath)
        print(f"Excel test report successfully generated and saved to {os.path.abspath(self.filepath)}")
